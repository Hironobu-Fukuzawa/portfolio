import calendar
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "2025_June_to_Dec"

months = range(6, 13)  # 6月〜12月
year = 2025

events = {
    (6, 15): "BCA Building Control Act",
    (6, 16): "Building Control Regulations",
    (6, 22): "BCS Part 1–3 (Use, Structure)",
    (6, 23): "BCS Part 4–6 (Means of Escape)",
    (6, 29): "BCS Part 7–9 (Fire, Safety)",
    (6, 30): "ทบทวน BCA + BCS",
    (7, 6): "Fire Code – Part 1",
    (7, 7): "Fire Code – Part 2",
    (7, 13): "Fire Code – Part 3",
    (7, 14): "ทบทวน Fire Code",
    (7, 20): "Planning Act – Overview",
    (7, 21): "URA Development Guidelines (Residential)",
    (7, 27): "URA Development Guidelines (Non-Residential)",
    (7, 28): "Architect’s Act & Roles",
    (8, 3): "Architect’s Rules + Registration",
    (8, 4): "Code on Accessibility",
    (8, 10): "Universal Design + Green Mark",
    (8, 11): "ทบทวนรวมทุกหมวด",
    (8, 17): "รวม Quiz สั้นหัวข้อหลัก",
    (8, 18): "สรุป Mind Map",
    (8, 24): "Weekly Mock Test (Self-made)",
    (8, 25): "รวมสรุป BCS + Fire Code",
    (8, 31): "รวมสรุป Architect’s Act",
    (9, 1): "แบบฝึกหัดปี 2010–2011",
    (9, 7): "ข้อสอบ 2012 + Review",
    (9, 8): "ข้อสอบ 2013 + Review",
    (9, 14): "ข้อสอบ 2014 + Review",
    (9, 15): "ข้อสอบ 2015 + Review",
    (9, 21): "ข้อสอบ 2016 + Review",
    (9, 22): "ข้อสอบ 2017 + Review",
    (9, 28): "ทบทวน BCS & Fire Code",
    (9, 29): "Checklist จุดอ่อน",
    (10, 5): "ข้อสอบ 2018 + Review",
    (10, 6): "ข้อสอบ 2019 + Review",
    (10, 12): "ข้อสอบ 2020 + Review",
    (10, 13): "ข้อสอบ 2021 + Review",
    (10, 19): "ข้อสอบ 2022 + Review",
    (10, 20): "ข้อสอบ 2023 + Review",
    (10, 26): "ข้อสอบ 2024 (ล่าสุด)",
    (10, 27): "สรุป Mind Map สุดท้าย",
    (11, 2): "ทบทวนรวมทุกจุด / Mock Simulation",
}

row_start = 1

# 枠線スタイル
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# イベントありのセル用色
event_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

for month in months:
    # 年月タイトル
    ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=7)
    title_cell = ws.cell(row=row_start, column=1, value=f"{year} {calendar.month_name[month]}")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(size=16, bold=True)
    row_start += 1

    # 曜日ヘッダー（Mon〜Sun）
    weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for col, day_name in enumerate(weekdays, start=1):
        cell = ws.cell(row=row_start, column=col, value=day_name)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.border = thin_border
    row_start += 1

    # カレンダーの日付
    cal = calendar.monthcalendar(year, month)
    for week in cal:
        for col, day in enumerate(week, start=1):
            cell = ws.cell(row=row_start, column=col)
            if day == 0:
                cell.value = ""
            else:
                if (month, day) in events:
                    cell.value = f"{day}\n{events[(month, day)]}"
                    cell.fill = event_fill
                else:
                    cell.value = f"{day}"
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            cell.border = thin_border
        row_start += 1
    row_start += 1  # 月ごとのスペース行

# 正方形に近づけるように調整
for col in range(1, 8):
    ws.column_dimensions[chr(64 + col)].width = 20

for row in range(1, row_start):
    ws.row_dimensions[row].height = 60

# 保存
file_path = "2025_June_to_December_Calendar_EventsOnlyColored.xlsx"
wb.save(file_path)
print(f"Excel file saved: {file_path}")
